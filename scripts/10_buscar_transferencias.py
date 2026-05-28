"""
Script 10 — Buscar certificados de transferencia bancaria (TT) por folio.

Para cada folio en DISTRIBUCIÓN CARGA VB, busca el archivo correcto en
FFOIP 2022 (REGIONES)/{region}/10. Egresos (TT)/ usando patrones por región,
y actualiza voucher_pdf en master_subida.xlsx.

Patrones por región:
  ARICA:      '{folio} {nombre}.pdf'
  ATACAMA:    'Certificado Bancario - FOLIO {folio}.pdf' (case-insensitive)
  O'HIGGINS:  '{folio} - {nombre}.pdf'
  ÑUBLE:      '{folio} {nombre}.pdf'
  AYSÉN:      '{seq} {folio} {nombre}.pdf'  (folio es el 2° token)
  RM:         '{rut}.pdf'  (busca en carpeta raíz y Nacionales/)
  VALPARAÍSO: sin archivos disponibles

Uso:
    python scripts/10_buscar_transferencias.py           # análisis, no modifica nada
    python scripts/10_buscar_transferencias.py --aplicar  # actualiza master_subida.xlsx
"""

import argparse
import re
from pathlib import Path

import openpyxl
import pandas as pd

MASTER   = Path("data/master_subida.xlsx")
DIST     = Path("data/DISTRIBUCIÓN CARGA VB.xlsx")

EGRESOS_DIR: dict[str, Path | None] = {
    "ARICA":       REGIONES / "03. Arica y Parinacota" / "10. Egresos (TT)",
    "ATACAMA":     REGIONES / "06. Atacama"             / "10. Egresos (TT)",
    "O'HIGGINS":   REGIONES / "09. O'Higgins"           / "10. Egresos (TT)",
    "ÑUBLE":       REGIONES / "11. Ñuble"               / "10. Egresos (TT)",
    "AYSEN":       REGIONES / "16. Aysén"               / "10. Egresos (TT)",
    "RM":          REGIONES / "18. RM"                  / "10. Egresos (TT)",
    "VALPARAÍSO":  None,
    "VALAPARAISO": None,
}

# Palabras clave que identifican documentos NO válidos como TT bancaria
# "recepci" → recepción/recepcion  |  "recurso" → Recursos/recursos (recepción abreviada)
_NO_TT = re.compile(r"recepci|recurso|sigfe", re.IGNORECASE)


def _es_pdf_tt(nombre: str) -> bool:
    return nombre.lower().endswith(".pdf") and not _NO_TT.search(nombre)


def _buscar_arica(folio: str, folder: Path) -> Path | None:
    # Arica structure: 10. Egresos (TT)/{folio} {org}/ or REG {folio} {org}/
    # Folio is always in the DIRECTORY name; file name inside is inconsistent.
    pat_dir = re.compile(r"^(?:REG\s+)?" + re.escape(folio) + r"\b", re.IGNORECASE)
    for subdir in folder.iterdir():
        if not subdir.is_dir() or not pat_dir.match(subdir.name):
            continue
        for f in sorted(subdir.iterdir()):
            if _es_pdf_tt(f.name):
                return f
    return None


def _buscar_atacama(folio: str, folder: Path) -> Path | None:
    pat = re.compile(r"FOLIO\s+" + re.escape(folio) + r"\b", re.IGNORECASE)
    for f in folder.iterdir():
        if pat.search(f.name) and _es_pdf_tt(f.name):
            return f
    return None


def _buscar_ohiggins(folio: str, folder: Path) -> Path | None:
    pat = re.compile(r"^" + re.escape(folio) + r"\s*[-–]\s*", re.IGNORECASE)
    for f in folder.iterdir():
        if pat.match(f.name) and _es_pdf_tt(f.name):
            return f
    return None


def _buscar_nuble(folio: str, folder: Path) -> Path | None:
    pat = re.compile(r"^" + re.escape(folio) + r"\s", re.IGNORECASE)
    for f in folder.iterdir():
        if pat.match(f.name) and _es_pdf_tt(f.name):
            return f
    return None


def _buscar_aysen(folio: str, folder: Path) -> Path | None:
    # Formato: "{seq} {folio} {nombre}.pdf" — folio es el segundo token
    pat = re.compile(r"^\d+\s+" + re.escape(folio) + r"\s", re.IGNORECASE)
    for f in folder.iterdir():
        if pat.match(f.name) and _es_pdf_tt(f.name):
            return f
    return None


def _buscar_rm(rut: str, folder: Path) -> Path | None:
    if not rut or rut.lower() == "nan":
        return None
    nombre = f"{rut}.pdf"
    for carpeta in [folder, folder / "Nacionales"]:
        candidato = carpeta / nombre
        if candidato.exists():
            return candidato
    return None


_BUSCADORES = {
    "ARICA":     _buscar_arica,
    "ATACAMA":   _buscar_atacama,
    "O'HIGGINS": _buscar_ohiggins,
    "ÑUBLE":     _buscar_nuble,
    "AYSEN":     _buscar_aysen,
}


def buscar_tt(folio: str, region: str, rut: str) -> Path | None:
    """Devuelve el Path del archivo TT, o None si no se encuentra."""
    key = region.upper().strip()
    # Normalizar variantes de Valparaíso
    if "VALPARA" in key:
        return None

    folder = EGRESOS_DIR.get(key)
    if folder is None:
        return None
    if not folder.exists():
        return None

    if key == "RM":
        return _buscar_rm(rut, folder)

    buscador = _BUSCADORES.get(key)
    if buscador is None:
        return None

    return buscador(folio, folder)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--aplicar", action="store_true",
                        help="Actualizar voucher_pdf en master_subida.xlsx")
    args = parser.parse_args()

    master = pd.read_excel(MASTER)
    dist   = pd.read_excel(DIST)
    folios_dist = set(dist["ID Convenio"].astype(str).str.strip())

    df = master[master["folio"].astype(str).isin(folios_dist)].copy()

    encontrado:    list[dict] = []
    no_encontrado: list[dict] = []

    for _, row in df.iterrows():
        folio  = str(row["folio"])
        region = str(row.get("region", "")).strip().upper()
        rut    = str(row.get("rut", "")).strip()
        razon  = str(row.get("razon_social", ""))

        tt = buscar_tt(folio, region, rut)

        if tt:
            encontrado.append({
                "folio": folio, "region": region,
                "razon_social": razon, "archivo": str(tt),
            })
        else:
            no_encontrado.append({
                "folio": folio, "region": region, "razon_social": razon,
            })

    # ── Consola ───────────────────────────────────────────────────────────────
    print(f"\n{'='*65}")
    print(f"  BÚSQUEDA DE TRANSFERENCIAS (TT)")
    print(f"{'='*65}")
    print(f"\n✅  Encontrados:    {len(encontrado)}")
    print(f"🔴  No encontrados: {len(no_encontrado)}")

    print("\n── ENCONTRADOS POR REGIÓN ──")
    por_reg: dict[str, int] = {}
    for r in encontrado:
        por_reg[r["region"]] = por_reg.get(r["region"], 0) + 1
    for reg in sorted(por_reg):
        print(f"  {reg:<15} {por_reg[reg]:>3}")

    if no_encontrado:
        print("\n── NO ENCONTRADOS POR REGIÓN ──")
        no_reg: dict[str, list] = {}
        for r in no_encontrado:
            no_reg.setdefault(r["region"], []).append(r)
        for reg in sorted(no_reg):
            print(f"\n  [{reg}] ({len(no_reg[reg])} folios)")
            for r in no_reg[reg]:
                print(f"    {r['folio']} — {r['razon_social'][:50]}")

    # ── Actualizar master ─────────────────────────────────────────────────────
    if args.aplicar:
        tt_map = {r["folio"]: r["archivo"] for r in encontrado}

        wb = openpyxl.load_workbook(MASTER)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        col_folio   = headers.index("folio") + 1
        col_voucher = headers.index("voucher_pdf") + 1

        updated = 0
        for row in ws.iter_rows(min_row=2):
            fval = str(row[col_folio - 1].value).strip() if row[col_folio - 1].value else ""
            if fval in tt_map:
                row[col_voucher - 1].value = tt_map[fval]
                updated += 1

        wb.save(MASTER)
        print(f"\n✅  Master actualizado: {updated} filas con voucher_pdf apuntando al TT correcto")
        print(f"    → {MASTER}")
        print()
        print("  Próximo paso:")
        print("  python scripts/04_subir_documentos.py --solo-transferencias")
    else:
        print(f"\n  → Para aplicar al master:")
        print(f"    python scripts/10_buscar_transferencias.py --aplicar")
    print()


if __name__ == "__main__":
    main()
