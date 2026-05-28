"""
Script 03 v3: Armar master de subida (estructura uniforme)
===========================================================
Escanea directamente las carpetas por folio según la estructura
establecida en CRITERIOS.md. No depende de inventario_disco.csv.

Estructura esperada:
  Archivos/{NN}. {REGION} - FFOIP/{folio} - {RAZON_SOCIAL}/
    Convenio   - {folio}.pdf
    Resolución - {folio}.pdf
    Egreso     - {folio}.pdf
    Rendición  - {folio}.pdf
    Garantía   - {folio}.pdf  (opcional)

  Colaboradores del Estado/{año}/
    certificado_{RUT_SIN_PUNTOS}_{DV}.pdf

Uso:
    python scripts/03_armar_master.py
"""

import csv
import re
import unicodedata
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ====== CONFIGURACIÓN ======
BASE_DIR           = Path("/home/bgcorrea/personal/workspace/caigg")
BASE_ARCHIVOS      = BASE_DIR / "Archivos"
BASE_COLABORADORES = BASE_DIR / "Colaboradores del Estado"

EXCEL_INPUT = "data/Libro1.xlsx"
CSV_RUT     = "logs/rut_por_folio_fixed.csv"
XLSX_OUTPUT = "data/master_subida.xlsx"
CSV_OUTPUT  = "logs/master_subida.csv"

EXT_PDF_SOLO = {".pdf"}
EXT_PDF_IMG  = {".pdf", ".jpg", ".jpeg", ".png"}

REGION_DIRS = {
    "ARICA":      "01. ARICA - FFOIP",
    "ATACAMA":    "04. ATACAMA - FFOIP",
    "VALPARAISO": "06. VALPARAÍSO - FFOIP",
    "RM":         "07. METROPOLITANA - FFOIP",
    "OHIGGINS":   "08. O'HIGGINS - FFOIP",
    "NUBLE":      "10. ÑUBLE - FFOIP",
    "AYSEN":      "15. AYSEN - FFOIP",
}

# Para regiones donde los folios no están en la raíz sino en una subcarpeta
REGION_SUBFOLDER = {
    "RM": "07. Firma de Convenios",
}

NOMBRE_ESTANDAR = {
    "convenio":   "Convenio",
    "resolucion": "Resolución",
    "egreso":     "Egreso",
    "rendicion":  "Rendición",
    "garantia":   "Garantía",
}

KEYWORDS = {
    "convenio":   ["convenio"],
    "resolucion": ["resolución", "resolucion", "res. adj", "rex", "exenta",
                   "adjudicación", "adjudicacion"],
    "egreso":     ["egreso", "transferencia", "recepción de recursos",
                   "recepcion de recursos", "voucher", "certificado bancario",
                   "certificado recepción", "certificado recepcion"],
    "rendicion":  ["rendición", "rendicion", "cfc", "fiel cumplimiento", "memo daf"],
    "garantia":   ["garantía", "garantia", "letra de cambio"],
}

EXTENSIONES_VALIDAS = {
    "convenio":   EXT_PDF_SOLO,
    "resolucion": EXT_PDF_SOLO,
    "egreso":     EXT_PDF_IMG,
    "rendicion":  EXT_PDF_SOLO,
    "garantia":   EXT_PDF_IMG,
}

FOLIO_PATRON = re.compile(r"^(\d{5,6})\s*-")


# ====== UTILIDADES ======

def normalizar(texto: str) -> str:
    if not texto:
        return ""
    nfd = unicodedata.normalize("NFD", texto)
    sin_tildes = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", sin_tildes.upper().strip())


def region_canonica(region: str) -> str:
    r = normalizar(region)
    if "ARICA" in r:        return "ARICA"
    if "ATACAMA" in r:      return "ATACAMA"
    if "VALPARAISO" in r:   return "VALPARAISO"
    if "HIGGINS" in r:      return "OHIGGINS"
    if "NUBLE" in r:        return "NUBLE"
    if "AYSEN" in r:        return "AYSEN"
    if r == "RM" or "METROPOLITANA" in r: return "RM"
    return r


def detectar_tipo(nombre: str) -> str | None:
    nombre_lower = nombre.lower()
    for tipo, prefijo in NOMBRE_ESTANDAR.items():
        if re.match(rf"^{re.escape(prefijo.lower())}\s*-\s*\d+", nombre_lower):
            return tipo
    for tipo, kws in KEYWORDS.items():
        if any(kw in nombre_lower for kw in kws):
            return tipo
    return None


def rut_sin_dv(rut: str) -> str:
    if not rut:
        return ""
    limpio = rut.replace(".", "").replace(" ", "")
    return limpio.split("-")[0] if "-" in limpio else limpio


# ====== CARGA DE DATOS ======

def cargar_excel() -> dict[str, dict]:
    wb = openpyxl.load_workbook(EXCEL_INPUT, data_only=True)
    ws = wb.active
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[2]:
            continue
        folio = str(row[2]).strip()
        out[folio] = {
            "fondo":             row[0] or "",
            "ano":               int(row[1]) if row[1] else 2022,
            "region":            (row[5] or "").strip().upper().replace("VALAPARAISO", "VALPARAÍSO"),
            "alcance":           row[6] or "",
            "acto_admin_tipo":   row[8] or "",
            "acto_admin_numero": str(row[9]).strip() if row[9] else "",
            "acto_admin_fecha":  row[10],
        }
    return out


def cargar_rut_csv() -> dict[str, dict]:
    out = {}
    with open(CSV_RUT, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            out[r["folio"]] = r
    return out


# ====== BÚSQUEDA DE ARCHIVOS ======

def encontrar_carpeta_folio(folio: str, region_dir: Path) -> Path | None:
    if not region_dir or not region_dir.exists():
        return None
    for sub in region_dir.iterdir():
        if not sub.is_dir():
            continue
        m = FOLIO_PATRON.match(sub.name)
        if m and m.group(1) == folio:
            return sub
    return None


def buscar_en_carpeta(folio: str, carpeta: Path) -> dict[str, tuple[str, str]]:
    """
    Escanea la carpeta del folio y devuelve {tipo: (ruta, confianza)}.
    Prioriza archivos con nombre estándar. Si hay ambigüedad sin canónico, deja vacío.
    """
    if not carpeta or not carpeta.exists():
        return {}

    por_tipo: dict[str, list[tuple[Path, bool]]] = {t: [] for t in NOMBRE_ESTANDAR}

    for f in sorted(carpeta.iterdir()):
        if not f.is_file():
            continue
        ext = f.suffix.lower()
        if ext not in {".pdf", ".jpg", ".jpeg", ".png", ".docx"}:
            continue
        tipo = detectar_tipo(f.name)
        if tipo not in por_tipo:
            continue
        estandar = bool(re.match(
            rf"^{re.escape(NOMBRE_ESTANDAR[tipo].lower())}\s*-\s*{folio}",
            f.name.lower()
        ))
        por_tipo[tipo].append((f, estandar))

    resultado = {}
    for tipo, archivos in por_tipo.items():
        if not archivos:
            continue
        estandares = [item for item in archivos if item[1]]
        if len(estandares) == 1:
            f, _ = estandares[0]
            if f.suffix.lower() in EXTENSIONES_VALIDAS[tipo]:
                resultado[tipo] = (str(f), "ALTA")
        elif len(archivos) == 1:
            f, estandar = archivos[0]
            if f.suffix.lower() in EXTENSIONES_VALIDAS[tipo]:
                resultado[tipo] = (str(f), "ALTA" if estandar else "MEDIA")
        # múltiples sin canónico: ambigüedad — no asignar

    return resultado


def buscar_certificado(rut: str, ano: int) -> tuple[str, str]:
    if not rut:
        return "", ""
    rut_clean = rut_sin_dv(rut)
    if not rut_clean:
        return "", ""

    anos = [str(ano)] + [str(y) for y in range(2022, 2026) if str(y) != str(ano)]
    for a_str in anos:
        carpeta_ano = BASE_COLABORADORES / a_str
        if not carpeta_ano.exists():
            continue
        for f in carpeta_ano.iterdir():
            if not f.is_file() or f.suffix.lower() != ".pdf":
                continue
            nombre_clean = f.name.replace(".", "").replace("-", "").replace(" ", "")
            if rut_clean in nombre_clean:
                return str(f), "ALTA"

    return "", ""


# ====== PROCESAMIENTO POR FOLIO ======

def procesar_folio(folio: str, datos_excel: dict, datos_sira: dict) -> dict:
    ano    = datos_excel.get("ano", 2022)
    region = datos_excel.get("region", "")
    rut    = datos_sira.get("rut", "")
    razon  = datos_sira.get("razon_social", "")
    estado = datos_sira.get("estado_sira", "")
    cuotas = datos_sira.get("cuotas_declaradas", "1")

    region_c       = region_canonica(region)
    region_dirname = REGION_DIRS.get(region_c, "")
    region_dir     = BASE_ARCHIVOS / region_dirname if region_dirname else None

    # Algunas regiones tienen sus carpetas de folio en una subcarpeta
    subfolder = REGION_SUBFOLDER.get(region_c)
    if region_dir and subfolder:
        region_dir = region_dir / subfolder

    carpeta = encontrar_carpeta_folio(folio, region_dir) if region_dir else None
    docs    = buscar_en_carpeta(folio, carpeta)

    def get(tipo):
        return docs.get(tipo, ("", ""))

    convenio_path,  conf_convenio   = get("convenio")
    acto_path,      conf_acto       = get("resolucion")
    voucher_path,   conf_voucher    = get("egreso")
    rendicion_path, conf_rendicion  = get("rendicion")
    garantia_path,  conf_garantia   = get("garantia")
    cert_path,      conf_cert       = buscar_certificado(rut, ano)

    n_encontrados = sum(bool(p) for p in [
        convenio_path, acto_path, cert_path, voucher_path, rendicion_path, garantia_path
    ])

    observaciones = []
    if estado == "Enviado":
        observaciones.append("YA ENVIADO - SKIP")
    if not carpeta:
        observaciones.append("SIN_CARPETA")
    if not convenio_path:
        observaciones.append("Sin convenio")
    if not acto_path:
        observaciones.append("Sin resolución")
    if not cert_path:
        observaciones.append("Sin certificado")
    if not voucher_path:
        observaciones.append("Sin egreso")
    if not rendicion_path:
        observaciones.append("Sin rendición")

    return {
        "folio":               folio,
        "ano":                 ano,
        "fondo":               datos_excel.get("fondo", ""),
        "region":              region,
        "alcance":             datos_excel.get("alcance", ""),
        "razon_social":        razon,
        "nombre_fantasia":     datos_sira.get("nombre_fantasia", ""),
        "rut":                 rut,
        "estado_sira":         estado,
        "cuotas_declaradas":   cuotas,
        "monto_garantia":      datos_sira.get("monto_garantia", ""),
        "acto_admin_numero":   datos_excel.get("acto_admin_numero", ""),
        "carpeta_folio":       str(carpeta) if carpeta else "",
        "convenio_pdf":        convenio_path,
        "acto_admin_pdf":      acto_path,
        "certificado_pdf":     cert_path,
        "voucher_pdf":         voucher_path,
        "rendicion_pdf":       rendicion_path,
        "garantia_pdf":        garantia_path,
        "confianza_convenio":     conf_convenio,
        "confianza_acto":         conf_acto,
        "confianza_certificado":  conf_cert,
        "confianza_voucher":      conf_voucher,
        "confianza_rendicion":    conf_rendicion,
        "confianza_garantia":     conf_garantia,
        "n_archivos_encontrados": n_encontrados,
        "observaciones":       " | ".join(observaciones) if observaciones else "OK",
        "estado_carga":        "",
    }


# ====== ESCRITURA DEL EXCEL ======

CAMPOS = [
    "folio", "ano", "fondo", "region", "alcance",
    "razon_social", "nombre_fantasia", "rut",
    "estado_sira", "cuotas_declaradas", "monto_garantia", "acto_admin_numero",
    "carpeta_folio",
    "convenio_pdf", "acto_admin_pdf", "certificado_pdf",
    "voucher_pdf", "rendicion_pdf", "garantia_pdf",
    "confianza_convenio", "confianza_acto", "confianza_certificado",
    "confianza_voucher", "confianza_rendicion", "confianza_garantia",
    "n_archivos_encontrados", "observaciones", "estado_carga",
]


def escribir_excel(filas: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "master_subida"

    FILL_HEADER  = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    FONT_HEADER  = Font(color="FFFFFF", bold=True, size=10)
    FILL_ALTA    = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    FILL_MEDIA   = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    FILL_VACIO   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    FILL_ENVIADO = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    for col_idx, campo in enumerate(CAMPOS, 1):
        c = ws.cell(row=1, column=col_idx, value=campo)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center")

    CONF_PATH_PAIRS = [
        ("confianza_convenio",    "convenio_pdf"),
        ("confianza_acto",        "acto_admin_pdf"),
        ("confianza_certificado", "certificado_pdf"),
        ("confianza_voucher",     "voucher_pdf"),
        ("confianza_rendicion",   "rendicion_pdf"),
        ("confianza_garantia",    "garantia_pdf"),
    ]

    for row_idx, fila in enumerate(filas, 2):
        for col_idx, campo in enumerate(CAMPOS, 1):
            ws.cell(row=row_idx, column=col_idx, value=fila.get(campo, ""))

        for conf_campo, path_campo in CONF_PATH_PAIRS:
            col_conf = CAMPOS.index(conf_campo) + 1
            col_path = CAMPOS.index(path_campo) + 1
            valor = fila.get(conf_campo, "")
            if valor == "ALTA":
                ws.cell(row=row_idx, column=col_conf).fill = FILL_ALTA
                ws.cell(row=row_idx, column=col_path).fill = FILL_ALTA
            elif valor == "MEDIA":
                ws.cell(row=row_idx, column=col_conf).fill = FILL_MEDIA
                ws.cell(row=row_idx, column=col_path).fill = FILL_MEDIA
            elif valor == "":
                ws.cell(row=row_idx, column=col_path).fill = FILL_VACIO

        if fila.get("estado_sira") == "Enviado":
            for c in range(1, len(CAMPOS) + 1):
                ws.cell(row=row_idx, column=c).fill = FILL_ENVIADO

    anchos = {
        "folio": 8, "ano": 6, "fondo": 8, "region": 14, "alcance": 12,
        "razon_social": 45, "nombre_fantasia": 30, "rut": 12,
        "estado_sira": 11, "cuotas_declaradas": 8, "monto_garantia": 14,
        "acto_admin_numero": 14, "carpeta_folio": 70,
        "convenio_pdf": 70, "acto_admin_pdf": 70, "certificado_pdf": 70,
        "voucher_pdf": 70, "rendicion_pdf": 70, "garantia_pdf": 70,
        "confianza_convenio": 10, "confianza_acto": 10, "confianza_certificado": 10,
        "confianza_voucher": 10, "confianza_rendicion": 10, "confianza_garantia": 10,
        "n_archivos_encontrados": 8, "observaciones": 45, "estado_carga": 12,
    }
    for col_idx, campo in enumerate(CAMPOS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = anchos.get(campo, 15)
    ws.freeze_panes = "A2"

    # Hoja resumen
    ws_r = wb.create_sheet("resumen")
    ws_r["A1"] = "MÉTRICA"
    ws_r["B1"] = "VALOR"
    ws_r["A1"].font = Font(bold=True)
    ws_r["B1"].font = Font(bold=True)

    total       = len(filas)
    enviados    = sum(1 for f in filas if f.get("estado_sira") == "Enviado")
    procesables = total - enviados
    completos   = sum(1 for f in filas
                      if f.get("estado_sira") != "Enviado"
                      and f.get("convenio_pdf") and f.get("acto_admin_pdf")
                      and f.get("certificado_pdf") and f.get("voucher_pdf")
                      and f.get("rendicion_pdf"))

    ws_r.append(["Total folios", total])
    ws_r.append(["Ya enviados (skip)", enviados])
    ws_r.append(["Procesables", procesables])
    ws_r.append(["Con todos los docs obligatorios completos", completos])
    ws_r.append([""])
    ws_r.append(["COBERTURA POR REGIÓN:", ""])
    ws_r.append([""])

    fila_h = ws_r.max_row + 1
    for col_idx, h in enumerate(["Región", "Total", "Conv", "Acto", "Cert", "Egreso", "Rend", "Gar"], 1):
        ws_r.cell(row=fila_h, column=col_idx, value=h).font = Font(bold=True)

    por_region: dict[str, dict] = defaultdict(lambda: defaultdict(int))
    for f in filas:
        if f.get("estado_sira") == "Enviado":
            continue
        reg = region_canonica(f.get("region", "?"))
        por_region[reg]["total"] += 1
        if f.get("convenio_pdf"):    por_region[reg]["conv"] += 1
        if f.get("acto_admin_pdf"):  por_region[reg]["acto"] += 1
        if f.get("certificado_pdf"): por_region[reg]["cert"] += 1
        if f.get("voucher_pdf"):     por_region[reg]["vouch"] += 1
        if f.get("rendicion_pdf"):   por_region[reg]["rend"] += 1
        if f.get("garantia_pdf"):    por_region[reg]["gar"] += 1

    for reg, stats in sorted(por_region.items()):
        ws_r.append([reg, stats["total"], stats["conv"], stats["acto"],
                     stats["cert"], stats["vouch"], stats["rend"], stats["gar"]])

    ws_r.append([""])
    ws_r.append(["LEYENDA:", ""])
    fila_l = ws_r.max_row + 1
    ws_r.cell(row=fila_l, column=1, value="ALTA confianza").fill = FILL_ALTA; fila_l += 1
    ws_r.cell(row=fila_l, column=1, value="MEDIA confianza").fill = FILL_MEDIA; fila_l += 1
    ws_r.cell(row=fila_l, column=1, value="Sin asignar").fill = FILL_VACIO; fila_l += 1
    ws_r.cell(row=fila_l, column=1, value="Ya enviado").fill = FILL_ENVIADO

    ws_r.column_dimensions["A"].width = 45
    ws_r.column_dimensions["B"].width = 12

    Path(XLSX_OUTPUT).parent.mkdir(exist_ok=True)
    wb.save(XLSX_OUTPUT)

    with open(CSV_OUTPUT, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CAMPOS, extrasaction="ignore")
        w.writeheader()
        w.writerows(filas)


# ====== MAIN ======

def main():
    print("=" * 70)
    print("SCRIPT 03 v3: ARMAR MASTER (estructura uniforme por folio)")
    print("=" * 70)

    for path in [EXCEL_INPUT, CSV_RUT]:
        if not Path(path).exists():
            print(f"[ERROR] No existe: {path}")
            return

    print(f"\nLeyendo Excel: {EXCEL_INPUT}")
    excel_data = cargar_excel()
    print(f"  {len(excel_data)} folios")

    print(f"\nLeyendo RUTs SIRA: {CSV_RUT}")
    sira_data = cargar_rut_csv()
    print(f"  {len(sira_data)} folios")

    print(f"\nProcesando {len(excel_data)} folios...")
    filas = []
    sin_sira = []
    for folio in sorted(excel_data):
        if folio not in sira_data:
            sin_sira.append(folio)
            continue
        filas.append(procesar_folio(folio, excel_data[folio], sira_data[folio]))

    if sin_sira:
        print(f"  [WARN] {len(sin_sira)} folios sin datos SIRA: {sin_sira[:5]}")

    print("\n" + "=" * 70)
    print("COBERTURA POR REGIÓN")
    print("=" * 70)

    por_region: dict[str, dict] = defaultdict(lambda: defaultdict(int))
    for f in filas:
        if f.get("estado_sira") == "Enviado":
            continue
        reg = region_canonica(f.get("region", "?"))
        por_region[reg]["total"] += 1
        if f.get("convenio_pdf"):    por_region[reg]["conv"] += 1
        if f.get("acto_admin_pdf"):  por_region[reg]["acto"] += 1
        if f.get("certificado_pdf"): por_region[reg]["cert"] += 1
        if f.get("voucher_pdf"):     por_region[reg]["vouch"] += 1
        if f.get("rendicion_pdf"):   por_region[reg]["rend"] += 1
        if f.get("garantia_pdf"):    por_region[reg]["gar"] += 1

    print(f"\n{'REGION':12s} | {'TOT':>4} | {'CONV':>5} | {'ACTO':>5} | {'CERT':>5} | {'EGRESO':>6} | {'REND':>5} | {'GAR':>5}")
    print("-" * 78)
    for reg, stats in sorted(por_region.items()):
        print(f"{reg:12s} | {stats['total']:>4d} | {stats['conv']:>5d} | {stats['acto']:>5d} | "
              f"{stats['cert']:>5d} | {stats['vouch']:>6d} | {stats['rend']:>5d} | {stats['gar']:>5d}")

    enviados    = sum(1 for f in filas if f.get("estado_sira") == "Enviado")
    procesables = len(filas) - enviados
    completos   = sum(1 for f in filas
                      if f.get("estado_sira") != "Enviado"
                      and f.get("convenio_pdf") and f.get("acto_admin_pdf")
                      and f.get("certificado_pdf") and f.get("voucher_pdf")
                      and f.get("rendicion_pdf"))

    print(f"\nTotal: {len(filas)}  |  Enviados: {enviados}  |  Procesables: {procesables}  |  Completos: {completos}")

    print(f"\nEscribiendo {XLSX_OUTPUT}...")
    escribir_excel(filas)
    print(f"OK — también: {CSV_OUTPUT}")

    print("\n" + "=" * 70)
    print("PRÓXIMOS PASOS")
    print("=" * 70)
    print(f"1. Revisá la hoja 'resumen' en {XLSX_OUTPUT}")
    print(f"2. Completá los archivos faltantes según script 06")
    print(f"3. Cuando esté listo: python scripts/04_subir_documentos.py")


if __name__ == "__main__":
    main()
