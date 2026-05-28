"""
Script 13 — Distribuir documentos faltantes a master: Convenios y Transferencias.

Fuentes:
  VALPARAÍSO Transferencias : 10. Egresos (TT)/0{n}.- RESOLUCIÓN {n}/{folio} {org}/
  VALPARAÍSO Convenios      : 07. Firma de Convenios/RESOLUCIÓN 1  CONVENIOS FIRMADOS/*.pdf
                               07. Firma de Convenios/02.- DOCUMENTOS.../{folio} - {org}/
                               07. Firma de Convenios/02.- DOCUMENTOS.../*.pdf (sueltos)
  ATACAMA    Convenios       : 07. Firma de Convenios/FOLIO {folio}/Convenio/*.pdf
  O'HIGGINS  Convenios       : 07. Firma de Convenios/ (folio-named + keyword map)
  ARICA      Convenio        : 07. Firma de Convenios/{folio} {org}/ (solo folio 60226)

No copia archivos — actualiza voucher_pdf / convenio_pdf en master_subida.xlsx
apuntando directo al origen.

Uso:
    python scripts/13_distribuir_faltantes.py           # análisis dry-run
    python scripts/13_distribuir_faltantes.py --aplicar # actualiza master
"""

import argparse
import re
from pathlib import Path

import openpyxl
import pandas as pd

MASTER   = Path("data/master_subida.xlsx")
DIST     = Path("data/DISTRIBUCIÓN CARGA VB.xlsx")

# ── Exclusión de archivos que no son documentos válidos ───────────────────────
_NO_DOC = re.compile(r"recepci|recurso|sigfe", re.IGNORECASE)

def _es_pdf(nombre: str) -> bool:
    return nombre.lower().endswith(".pdf") and not _NO_DOC.search(nombre)


# ── O'HIGGINS — mapa keyword → folio ─────────────────────────────────────────
# Archivo en 07. Firma de Convenios/ → folio
OH_KEYWORD: dict[str, str] = {
    "AGRUPACION AJEDREZ OHIGGINS":              "54768",
    "WE FOLIL":                                  "54899",
    "AGRUPACION LAGO RAPEL":                     "55756",
    "JJVV UNION CARACOLES":                      "55783",
    "EMPRENDEDORES DONIHUANOS":                  "55841",
    "ECO CLUBES LIFAE":                          "55883",
    "JJVV LAS PARCELAS TROYA NORTE":             "56367",
    "JJVV LA CABANA":                            "56431",
    "UCAM PAREDONES":                            "56573",
    "JJVV LA PALMILLA":                          "56623",
    "UCAM SAN FERNANDO":                         "56873",
    "CDS MOVIMIENTO FUNCIONAL":                  "57020",
    "TREBOL DE MAR":                             "57059",
    "JJVV ARICA":                                "57087",
    "CAM JAIME FELDMAN":                         "57169",
    "CAM BERNARDO MUNOZ":                        "57220",
    "UNIVERZOO ANIMAL":                          "57735",
    "OBSERVATORIO BUEN TRATO":                   "57823",
    "ARTESANOS EL PROGRESO":                     "58117",
    "AGRUPACION SIEMPRE UNIDOS":                 "58205",
    "LOS MONITOS DE MARY":                       "59076",
    "UCAM OLIVAR":                               "59266",
    "HIJOS Y PADRES MIRANDO HACIA EL FUTURO":    "59388",
    "COMITE ROSARIO AL PROGRESO":                "59410",
    "UCOM GRANEROS":                             "59433",
    "UCAM MARCHIGUE":                            "59561",
    "UCAM MOSTAZAL":                             "59671",
    "JJVV LAS HIGUERAS":                         "60005",
    "JJVV BERNARDO RETAMAL":                     "60058",
}

# ── Valparaíso — mapa de sueltos en 02-DOCUMENTOS → folio ────────────────────
# Archivos sueltos cuyos nombres no empiezan con folio
VALPO_SUELTOS: dict[str, str] = {
    "cc teatro confin":              "54929",
    "centro juvenil juventud futuro": "54870",
    "circulo teatral":               "57647",
    "fund somos":                    "55767",
    "jjvv bernardo leigh":           "56542",
    "jjvv ramaditas parte baja":     "58082",
    "ong accion intercultural":      "57822",
}


# ══════════════════════════════════════════════════════════════════════════════
# Funciones de búsqueda por región / documento
# ══════════════════════════════════════════════════════════════════════════════

def buscar_valpo_tt(folio: str) -> Path | None:
    """
    VALPARAÍSO TT: 10. Egresos (TT)/0{n}.- RESOLUCIÓN {n}/{folio} {org}/ → primer PDF válido.
    """
    base = REGIONES / "08. Valparaíso" / "10. Egresos (TT)"
    pat = re.compile(r"^" + re.escape(folio) + r"\b", re.IGNORECASE)
    for res_dir in sorted(base.iterdir()):
        if not res_dir.is_dir():
            continue
        for org_dir in res_dir.iterdir():
            if not org_dir.is_dir() or not pat.match(org_dir.name):
                continue
            for f in sorted(org_dir.iterdir()):
                if _es_pdf(f.name):
                    return f
    return None


def buscar_valpo_convenio(folio: str) -> Path | None:
    """
    VALPARAÍSO Convenio: busca en dos fuentes:
      1. RESOLUCIÓN 1  CONVENIOS FIRMADOS/{folio} *.pdf
      2. 02.- DOCUMENTOS.../{folio} - {org}/ subfolders
      3. 02.- DOCUMENTOS.../*.pdf sueltos (keyword map)
    """
    firma = REGIONES / "08. Valparaíso" / "07. Firma de Convenios"
    pat = re.compile(r"^" + re.escape(folio) + r"\b", re.IGNORECASE)

    # 1. RESOLUCIÓN 1 CONVENIOS FIRMADOS — fichero directo
    res1 = firma / "RESOLUCIÓN 1  CONVENIOS FIRMADOS"
    if res1.exists():
        for f in res1.iterdir():
            if f.is_file() and pat.match(f.name) and _es_pdf(f.name):
                return f

    # 2. 02.- DOCUMENTOS subfolder  {folio} - {org} / *.pdf
    docs = firma / "02.- DOCUMENTOS FIRMA DE CONVENIOS ORGANIZACIONES"
    if docs.exists():
        for subdir in docs.iterdir():
            if subdir.is_dir() and pat.match(subdir.name):
                for f in sorted(subdir.iterdir()):
                    if f.is_file() and _es_pdf(f.name):
                        return f

    # 3. 02.- DOCUMENTOS — PDFs sueltos mapeados por keyword
    if docs.exists():
        for kw, f_folio in VALPO_SUELTOS.items():
            if f_folio != folio:
                continue
            for f in docs.iterdir():
                if f.is_file() and kw in f.name.lower() and _es_pdf(f.name):
                    return f

    return None


def buscar_atacama_convenio(folio: str) -> Path | None:
    """
    ATACAMA Convenio: 07. Firma de Convenios/FOLIO {folio}/Convenio/*.pdf
    """
    folder = REGIONES / "06. Atacama" / "07. Firma de Convenios" / f"FOLIO {folio}"
    convenio_dir = folder / "Convenio"
    if convenio_dir.exists():
        for f in sorted(convenio_dir.iterdir()):
            if f.is_file() and _es_pdf(f.name):
                return f
    # Fallback: any PDF directly in the FOLIO folder
    if folder.exists():
        for f in sorted(folder.iterdir()):
            if f.is_file() and _es_pdf(f.name):
                return f
    return None


def buscar_ohiggins_convenio(folio: str, razon_social: str) -> Path | None:
    """
    O'HIGGINS Convenio: 07. Firma de Convenios/
      - Archivo con folio al inicio: {folio} - {org}.pdf
      - Archivo por keyword map: OH_KEYWORD
    """
    folder = REGIONES / "09. O'Higgins" / "07. Firma de Convenios"
    if not folder.exists():
        return None

    # Folio-named files
    pat = re.compile(r"^" + re.escape(folio) + r"\b", re.IGNORECASE)
    for f in folder.iterdir():
        if f.is_file() and pat.match(f.name) and _es_pdf(f.name):
            return f

    # Keyword map
    mapped = OH_KEYWORD.get(folio)  # inverted below
    return None  # handled by caller via inverted map


def buscar_arica_convenio(folio: str) -> Path | None:
    """
    ARICA Convenio: 07. Firma de Convenios/{folio} {org}/ → primer PDF válido.
    """
    folder = REGIONES / "03. Arica y Parinacota" / "07. Firma de Convenios"
    if not folder.exists():
        return None
    pat = re.compile(r"^" + re.escape(folio) + r"\b", re.IGNORECASE)
    for subdir in folder.iterdir():
        if subdir.is_dir() and pat.match(subdir.name):
            for f in sorted(subdir.iterdir()):
                if f.is_file() and _es_pdf(f.name):
                    return f
    return None


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--aplicar", action="store_true",
                        help="Actualizar master_subida.xlsx")
    args = parser.parse_args()

    master = pd.read_excel(MASTER)
    dist   = pd.read_excel(DIST)
    folios_dist = set(dist["ID Convenio"].astype(str).str.strip())

    # Invertir keyword map para O'Higgins: folio → filename_path
    oh_file_map: dict[str, Path] = {}
    oh_folder = REGIONES / "09. O'Higgins" / "07. Firma de Convenios"
    if oh_folder.exists():
        for f in oh_folder.iterdir():
            if not f.is_file() or not _es_pdf(f.name):
                continue
            stem_upper = f.stem.upper().strip()
            for kw, fol in OH_KEYWORD.items():
                if kw in stem_upper or stem_upper in kw:
                    oh_file_map[fol] = f
                    break

    # folio → {campo → path}  (un folio puede tener varios campos encontrados)
    encontrado:    dict[str, dict[str, str]] = {}
    no_encontrado: dict[str, list] = {}   # region → list of {folio, doc}

    def registrar(folio, region, doc, path):
        if path and path.exists():
            encontrado.setdefault(folio, {})
            encontrado[folio][doc] = str(path)
            encontrado[folio].setdefault("region", region)
        else:
            no_encontrado.setdefault(region, []).append({"folio": folio, "doc": doc})

    for _, row in master.iterrows():
        folio  = str(row["folio"])
        if folio not in folios_dist:
            continue

        region = str(row.get("region", "")).strip().upper()
        razon  = str(row.get("razon_social", "")).strip()

        def _falta(campo):
            v = str(row.get(campo, "")).strip()
            return not v or v == "nan" or not Path(v).exists()

        # ── VALPARAÍSO ──────────────────────────────────────────────────────
        if "VALPARA" in region:
            # TT: siempre buscar — el voucher anterior puede ser "Egreso" (recepción), no real TT
            registrar(folio, region, "voucher_pdf", buscar_valpo_tt(folio))
            if _falta("convenio_pdf"):
                registrar(folio, region + "_CONV", "convenio_pdf",
                          buscar_valpo_convenio(folio))

        # ── ATACAMA ─────────────────────────────────────────────────────────
        elif region == "ATACAMA":
            if _falta("convenio_pdf"):
                registrar(folio, region, "convenio_pdf",
                          buscar_atacama_convenio(folio))

        # ── O'HIGGINS ───────────────────────────────────────────────────────
        elif region == "O'HIGGINS":
            if _falta("convenio_pdf"):
                # Try folio-named file first, then keyword map
                path = buscar_ohiggins_convenio(folio, razon)
                if path is None:
                    path = oh_file_map.get(folio)
                registrar(folio, region, "convenio_pdf", path)

        # ── ARICA ───────────────────────────────────────────────────────────
        elif region == "ARICA":
            if _falta("convenio_pdf"):
                registrar(folio, region, "convenio_pdf",
                          buscar_arica_convenio(folio))

    # ── Reporte consola ───────────────────────────────────────────────────────
    print(f"\n{'='*65}")
    print(f"  DISTRIBUCIÓN DE DOCUMENTOS FALTANTES")
    print(f"{'='*65}")
    print(f"\n✅  Encontrados: {len(encontrado)}")
    total_nf = sum(len(v) for v in no_encontrado.values())
    print(f"🔴  No encontrados: {total_nf}")

    # Desglose por región+doc
    from collections import Counter
    by_reg: Counter = Counter()
    for info in encontrado.values():
        region_val = info.get("region", "")
        for campo in info:
            if campo == "region":
                continue
            by_reg[(region_val, campo)] += 1
    print("\n── ENCONTRADOS ──")
    for (reg, doc), cnt in sorted(by_reg.items()):
        doc_corto = doc.replace("_pdf", "").replace("voucher", "Transferencias").replace("convenio", "Convenio")
        print(f"  {reg:<20} {doc_corto:<16} {cnt:>3}")

    print("\n── NO ENCONTRADOS ──")
    for reg in sorted(no_encontrado):
        items = no_encontrado[reg]
        print(f"\n  [{reg}] ({len(items)})")
        for r in items:
            doc_c = r["doc"].replace("_pdf","").replace("voucher","Transferencias").replace("convenio","Convenio")
            print(f"    {r['folio']} — {doc_c}")

    # ── Actualizar master ─────────────────────────────────────────────────────
    if args.aplicar:
        wb = openpyxl.load_workbook(MASTER)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        col_folio = headers.index("folio") + 1

        campo_cols = {}
        for campo in ["convenio_pdf", "voucher_pdf", "rendicion_pdf"]:
            if campo in headers:
                campo_cols[campo] = headers.index(campo) + 1

        updated = 0
        for row in ws.iter_rows(min_row=2):
            fval = str(row[col_folio - 1].value).strip() if row[col_folio - 1].value else ""
            if fval not in encontrado:
                continue
            for campo, path_val in encontrado[fval].items():
                if campo == "region":
                    continue
                if campo in campo_cols:
                    row[campo_cols[campo] - 1].value = path_val
                    updated += 1

        wb.save(MASTER)
        print(f"\n✅  Master actualizado: {updated} campos actualizados → {MASTER}")
        print()
        print("  Próximos pasos:")
        print("  1. python scripts/09_listos_para_enviar.py   # ver nuevos listos")
        print("  2. python scripts/04_subir_documentos.py     # subir nuevos docs")
    else:
        print(f"\n  → Para aplicar:")
        print(f"    python scripts/13_distribuir_faltantes.py --aplicar")
    print()


if __name__ == "__main__":
    main()
