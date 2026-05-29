"""
Analiza los archivos en ARCHIVOS-SUBIR y los cruza con Libro1.xlsx
para generar un reporte de mapeo propuesto antes de normalizar.

Salida: logs/mapeo_propuesto.xlsx
"""

import os
import re
import unicodedata
import openpyxl
from openpyxl.styles import PatternFill, Font
from pathlib import Path

BASE = Path(__file__).parent.parent
LIBRO1 = BASE / "Libro1.xlsx"
ARCHIVOS = BASE / "ARCHIVOS-SUBIR"
OUT = BASE / "logs" / "mapeo_propuesto.xlsx"

# ── helpers ──────────────────────────────────────────────────────────────────

def normalizar(texto):
    """Minúsculas, sin tildes, sin caracteres especiales."""
    texto = texto.lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^a-z0-9 ]", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def extraer_folio_del_nombre(nombre):
    """Busca un número de 5-6 dígitos en el nombre del archivo."""
    m = re.search(r'\b(\d{5,6})\b', nombre)
    if m:
        return int(m.group(1))
    # También captura F74253 (sin espacio)
    m = re.search(r'[Ff](\d{5,6})', nombre)
    if m:
        return int(m.group(1))
    return None


def score_similitud(clave, nombre_norm):
    """Devuelve cuántas palabras de clave aparecen en nombre_norm."""
    palabras = clave.split()
    hits = sum(1 for p in palabras if p in nombre_norm)
    return hits / len(palabras) if palabras else 0


# ── leer Libro1 ───────────────────────────────────────────────────────────────

def leer_libro1():
    wb = openpyxl.load_workbook(LIBRO1)
    ws = wb.active
    folios = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        folio = row[8]   # ID Convenio
        rut   = row[5]   # RUT
        nombre = row[11] # NOMBRE OOSS/MEDIO
        if folio:
            folios[int(folio)] = {"rut": str(rut), "nombre": str(nombre)}
    return folios


# ── mapeo de CONVENIOS ────────────────────────────────────────────────────────

CONVENIO_MANUAL = {
    # nombre_archivo_base (sin fecha y extensión) → folio
    "adrian escobar": 77492,
    "amor": 74926,
    "bahia on line firmado": 77482,
    "caleta higuerillas": 74253,
    "calle larga": 77550,
    "canal 2": 73739,
    "canal local": 72589,
    "congreso": 75758,
    "costa magazine": 77176,
    "crystal": 77412,
    "de la costa": 77181,
    "eclipse": 70905,
    "fabiola tello": 77337,
    "global cosmos": 77821,
    "humberto lopez  (crecer)firmado": 75516,
    "humberto lopez (crecer)firmado": 75516,
    "ilusion": 72476,
    "interferencia ok": 69805,
    "kudell": 76841,
    "la merced": 71801,
    "la quinta emprende": 74148,
    "latina": 73226,
    "lovely": 77125,
    "mata o te rapa nui": 73944,
    "montealegre": 77380,
    "multitudes perifericas": 77337,  # AMBIGUO: misma razón social que fabiola tello
    "pagina12": 77305,
    "para dios": 74649,
    "portales": 77562,
    # preludio: AMBIGUO — 5 archivos para 2 folios (70245 y 70236)
    "preludio fm ok27-07-2023-160839": 70245,  # propuesta provisional
    "preludio fm ok27-07-2023-161004": 70245,  # DUPLICADO probable
    "preludio fm27-07-2023-160808": 70236,
    "preludio ok27-07-2023-160052": 70236,
    "preludio27-07-2023-160006": 70236,
    "proa": 76889,
    "quilpue on line": 73951,
    "raudal": 72170,
    "rd aconcagua": 74989,
    "rd buena onda": 77437,   # AMBIGUO vs crystal/77412 — revisar
    "rd he": 71909,
    "rd quintay firmado": 74637,
    "rd somos": 77263,        # AMBIGUO vs somos la ligua
    "rd valparaiso": 77257,
    "somos la ligua": 77798,  # AMBIGUO — podría ser 77263
    "super andina": 70793,
    "tu opinas": 71206,
    "tv lanligua": 70037,
    "vistamar": 72699,
    # "amiga": None  ← terminado, no está en Libro1
}

CONVENIO_OBSERVACIONES = {
    77337: "AMBIGUO: dos archivos (fabiola tello + multitudes perifericas) → mismo folio. Conservar solo el correcto.",
    70245: "AMBIGUO: 5 archivos de preludio para 2 folios (70245 y 70236). Revisar cuál corresponde a cada uno.",
    70236: "AMBIGUO: ver folio 70245.",
    77412: "AMBIGUO: crystal ↔ buena onda. Cruzar con SIRA para confirmar.",
    77437: "AMBIGUO: rd buena onda asignado aquí pero egreso dice 'RADIO CRYSTAL'. Revisar.",
    77263: "AMBIGUO: rd somos vs somos la ligua — confirmar cuál es 77263 y cuál 77798.",
    77798: "AMBIGUO: somos la ligua vs rd somos — confirmar.",
}

FOLIOS_IGNORAR = {77693, 72374}  # fuera de Libro1 o TAC


# ── inventariar archivos fuente ───────────────────────────────────────────────

def clave_convenio(nombre_archivo):
    """Extrae la clave del nombre de un archivo de Convenio (sin fecha ni extensión)."""
    base = Path(nombre_archivo).stem
    # Eliminar sufijo de fecha tipo "27-07-2023-150711"
    base = re.sub(r'\d{2}-\d{2}-\d{4}-\d{6}$', '', base).strip()
    base = re.sub(r'\d{2}-\d{2}-\d{4}-\d{4}$', '', base).strip()
    return base.lower()


def procesar_convenios(folios_libro1):
    filas = []
    carpeta = ARCHIVOS / "Convenios"
    for f in sorted(os.listdir(carpeta)):
        if not f.endswith(".pdf"):
            continue
        ruta = str(carpeta / f)
        base_sin_fecha = clave_convenio(f)

        # Intentar mapeo directo
        folio = CONVENIO_MANUAL.get(base_sin_fecha)
        # Si no hay clave exacta, buscar clave parcial
        if folio is None:
            for k, v in CONVENIO_MANUAL.items():
                if base_sin_fecha.startswith(k) or k.startswith(base_sin_fecha):
                    folio = v
                    break

        if folio is None and "amiga" in base_sin_fecha:
            filas.append({
                "tipo": "CONVENIO", "archivo_origen": ruta,
                "folio": "SIN FOLIO", "razon_social": "NO EN LIBRO1 (convenio terminado)",
                "nombre_destino": "N/A", "confianza": "IGNORAR",
                "observacion": "Radio Amiga: convenio terminado, no está en Libro1"
            })
            continue

        if folio is None:
            filas.append({
                "tipo": "CONVENIO", "archivo_origen": ruta,
                "folio": "SIN FOLIO", "razon_social": "NO MAPEADO",
                "nombre_destino": "N/A", "confianza": "MANUAL",
                "observacion": f"No se encontró mapeo para: {base_sin_fecha}"
            })
            continue

        info = folios_libro1.get(folio, {})
        obs = CONVENIO_OBSERVACIONES.get(folio, "")
        confianza = "REVISAR" if obs else "ALTA"
        nombre_dest = f"Convenio - {folio}.pdf"
        filas.append({
            "tipo": "CONVENIO", "archivo_origen": ruta,
            "folio": folio, "razon_social": info.get("nombre", ""),
            "nombre_destino": nombre_dest, "confianza": confianza,
            "observacion": obs
        })
    return filas


def procesar_egresos(folios_libro1):
    filas = []
    carpeta = ARCHIVOS / "EGRESOS"
    EGRESO_MANUAL = {
        "AGRUPACION PENIEL LA LIGUA": 74649,
        "BAHIA ONLINE": 77482,
        "JJVV Villa Sto Domingo": 73305,
        "RADIO CRECER FM": 75516,
        "RADIO QUINTAY": 74637,
        "RADIO SOMOS": 77798,
        "RODRIGO BERNAL CACERES": 72589,
        "F74253 RADIO COMUN. CALETA HUGUERI": 74253,
        "F75516 RADIO CRECER": 75516,
    }
    EGRESO_IGNORAR = {"77693 RADIO 88.7"}
    for f in sorted(os.listdir(carpeta)):
        if not f.endswith(".pdf"):
            continue
        ruta = str(carpeta / f)
        stem = Path(f).stem

        # Ignorar fuera de Libro1
        if any(ig in stem for ig in ["77693", "RADIO CRECER FM"]):
            if "77693" in stem:
                filas.append({
                    "tipo": "EGRESO", "archivo_origen": ruta,
                    "folio": 77693, "razon_social": "NO EN LIBRO1 (Radio 88.7)",
                    "nombre_destino": "N/A", "confianza": "IGNORAR",
                    "observacion": "Folio 77693 no está en Libro1"
                })
                continue

        folio = extraer_folio_del_nombre(stem)
        if folio is None:
            # Buscar en mapeo manual
            for k, v in EGRESO_MANUAL.items():
                if normalizar(k) in normalizar(stem) or normalizar(stem) in normalizar(k):
                    folio = v
                    break

        if folio is None:
            filas.append({
                "tipo": "EGRESO", "archivo_origen": ruta,
                "folio": "SIN FOLIO", "razon_social": "NO MAPEADO",
                "nombre_destino": "N/A", "confianza": "MANUAL",
                "observacion": f"No se detectó folio en: {stem}"
            })
            continue

        info = folios_libro1.get(folio, {})
        en_libro = folio in folios_libro1
        filas.append({
            "tipo": "EGRESO", "archivo_origen": ruta,
            "folio": folio, "razon_social": info.get("nombre", "FUERA DE LIBRO1"),
            "nombre_destino": f"Egreso - {folio}.pdf",
            "confianza": "ALTA" if en_libro else "IGNORAR",
            "observacion": "" if en_libro else "Folio no en Libro1"
        })
    return filas


def procesar_rendiciones(folios_libro1):
    filas = []
    carpeta = ARCHIVOS / "GARANTIAS Y RENDICIONES" / "02. RENDICIONES"
    for f in sorted(os.listdir(carpeta)):
        if not f.endswith(".pdf"):
            continue
        ruta = str(carpeta / f)
        folio = extraer_folio_del_nombre(Path(f).stem)
        if folio is None:
            filas.append({
                "tipo": "RENDICION", "archivo_origen": ruta,
                "folio": "SIN FOLIO", "razon_social": "NO MAPEADO",
                "nombre_destino": "N/A", "confianza": "MANUAL",
                "observacion": f"No se detectó folio en: {f}"
            })
            continue
        if folio in FOLIOS_IGNORAR:
            info = "TAC" if folio == 72374 else "Fuera de Libro1"
            filas.append({
                "tipo": "RENDICION", "archivo_origen": ruta,
                "folio": folio, "razon_social": info,
                "nombre_destino": "N/A", "confianza": "IGNORAR",
                "observacion": f"Folio {folio} excluido"
            })
            continue
        info = folios_libro1.get(folio, {})
        en_libro = folio in folios_libro1
        filas.append({
            "tipo": "RENDICION", "archivo_origen": ruta,
            "folio": folio, "razon_social": info.get("nombre", "FUERA DE LIBRO1"),
            "nombre_destino": f"Rendición - {folio}.pdf",
            "confianza": "ALTA" if en_libro else "IGNORAR",
            "observacion": "" if en_libro else "Folio no en Libro1"
        })
    return filas


def procesar_garantias(folios_libro1):
    """Mapeo fuzzy por razón social."""
    filas = []
    carpeta = ARCHIVOS / "GARANTIAS Y RENDICIONES" / "01. GARANTÍAS"

    # Construir índice normalizado de Libro1
    indice = {normalizar(v["nombre"]): k for k, v in folios_libro1.items()}

    GARANTIA_MANUAL = {
        "88.7": None,                          # folio 77693 — fuera de Libro1
        "AGRUACIÓN PENIEL LA LIGUA": 74649,
        "AGRUPACION SOCIAL, CULTURAL Y COMUNITARIA": 69805,
        "BOX COMUNICACIONES SPA": 77798,
        "CANAL LOCAL": 72589,
        "CENTRO CULTURAL REMA COMUNICACIONES DE EL QUISCO": None,  # no en Libro1
        "CENTRO JUVENIL CULTURAL SOCIAL Y DE COMUNICACIONES RATEM": 71876,
        "COMUNICACIONES ACONCAGUA SPA": 77482,
        "COMUNICACIONES CONGRESO SPA": 75758,
        "COMUNICACIONES JULIO HARDOY BAYLAUCQ": None,   # no en Libro1
        "COMUNICACIONES PACIFICO SPA": 77257,
        "ESTUDIO TV LA LIGUA": 70037,
        "FABIOLA HERRERA LEIVA": None,         # no en Libro1
        "HERNAN PULGAR AGUILERA SPA": 71909,
        "HUMBERTO LOPEZ VERGARA": 75516,
        "JUNTA DE VECINOS CALETA HIGUERILLAS": 74253,
        "KUDELL TV": 76841,
        "LEONARDO PAKARATI": 73542,
        "LOVELY FM": 77125,
        "MANUEL DIAZ VILLAGRAN": 74919,
        "MIGUEL ANGEL JARA MALDONADO": None,   # no en Libro1
        "ORGANIZACION PARROQUIA NUESTRA SEÑORA DE LA MERCED": 71801,
        "PERIODICO DE LA COSTA": 77181,
        "PRELUDIO COMUNICACIONES SPA": 70236,
        "QUILPUE ONLINE": 73951,
        "RADIO CALLE LARGA": 77550,
        "RADIO ECLIPSE": 70905,
        "RADIO PORTALES": 77562,
        "RADIO RECREO SPA": 77003,
        "RADIO SOMOS": 77798,                  # AMBIGUO — misma razón social que BOX COMUNICACIONES
        "RADIO VISTAMAR": 72699,
        "RADIODIFUSION ADRIAN ESCOBAR ARAVENA": 77492,
        "RADIODIFUSION VERONICA DEL CARMEN": 77821,
        "REVISTA COSTA MAGAZINE": 77176,
        "SEMINARIO PAGINA 12": 77305,
        "SOCIEDAD COMUNICACIONES RAUDAL LTDA": 72170,
        "SOCIEDAD DIFUSORA DE RADIO Y TV SAN ANTONI": 73739,
        "SOCIEDAD PUBLIEVENTOS LTDA": 72009,
        "SOCIEDAD RADIO ACONCAGUA LIMITADA": 74989,
        "SUPERANDINA": 70793,
        "TU OPINAS.CL": 71206,
        "UNCO OLMUE": None,                    # no en Libro1 (Union Comunal Olmué = folio 76359?)
        "VICTORIA CALDERÓN QUEVEDO": None,     # no en Libro1
    }
    OBS_GARANTIA = {
        None: "No encontrado en Libro1 — revisar si corresponde a otro fondo",
        77798: "AMBIGUO: BOX COMUNICACIONES SPA y RADIO SOMOS podrían ser el mismo folio 77798",
        76359: "POSIBLE: UNCO OLMUE podría ser folio 76359 (UNION COMUNAL DE JUNTAS DE VECINOS DE OLMUE)",
    }

    for f in sorted(os.listdir(carpeta)):
        if not f.endswith(".pdf"):
            continue
        ruta = str(carpeta / f)
        stem = Path(f).stem

        folio = GARANTIA_MANUAL.get(stem)
        if folio is None and stem not in GARANTIA_MANUAL:
            # Búsqueda manual no registrada
            obs = f"Clave '{stem}' no en mapeo manual — agregar a GARANTIA_MANUAL"
            filas.append({
                "tipo": "GARANTIA", "archivo_origen": ruta,
                "folio": "SIN FOLIO", "razon_social": "NO MAPEADO",
                "nombre_destino": "N/A", "confianza": "MANUAL", "observacion": obs
            })
            continue

        info = folios_libro1.get(folio, {}) if folio else {}
        obs = OBS_GARANTIA.get(folio, "")
        if folio is None:
            obs = OBS_GARANTIA.get(None, "")

        confianza = "IGNORAR" if folio is None else ("REVISAR" if obs else "ALTA")
        filas.append({
            "tipo": "GARANTIA", "archivo_origen": ruta,
            "folio": folio if folio else "SIN FOLIO",
            "razon_social": info.get("nombre", "FUERA DE LIBRO1" if folio else "NO EN LIBRO1"),
            "nombre_destino": f"Garantía - {folio}.pdf" if folio else "N/A",
            "confianza": confianza, "observacion": obs
        })
    return filas


def procesar_resoluciones(folios_libro1):
    """Resolución principal → todos los folios de medios; específicas → su folio."""
    filas = []
    carpeta = ARCHIVOS / "RESOLUCIONES"

    # Folios de medios (los que tienen egreso detectado en Libro1)
    folios_medios = [
        69805, 70037, 70236, 70245, 70793, 70905, 71206, 71801, 71876, 71909,
        72009, 72170, 72476, 72589, 72699, 73226, 73542, 73739, 73944, 73951,
        74148, 74253, 74637, 74649, 74919, 74926, 74989, 75516, 75758, 76841,
        76889, 77003, 77125, 77176, 77181, 77257, 77263, 77305, 77337, 77380,
        77412, 77437, 77482, 77492, 77550, 77562, 77798, 77821
    ]
    folios_medios = [f for f in folios_medios if f in folios_libro1]

    RESOLUCION_ESPECIFICA = {
        "Res. Ex. 2379 Radio 88.7": None,              # folio 77693 fuera de Libro1
        "RES. TERMINO CONVENIO RADIO AMIGA": None,     # convenio terminado
        "Acuerdo 10899 Modifica beneficiarios FFMCS": "TODOS",
        "Acuerdo 10951 Modifica beneficiarios FFMCS": "TODOS",
    }
    RESOLUCION_PRINCIPAL = "2 APRUEBA CONVENIOS DE TRANSFERENCIA"

    for f in sorted(os.listdir(carpeta)):
        if not f.endswith(".pdf"):
            continue
        ruta = str(carpeta / f)
        stem = Path(f).stem

        if RESOLUCION_PRINCIPAL in stem:
            filas.append({
                "tipo": "RESOLUCION", "archivo_origen": ruta,
                "folio": "TODOS LOS MEDIOS",
                "razon_social": f"{len(folios_medios)} folios",
                "nombre_destino": "Resolución - {folio}.pdf (copia a cada carpeta)",
                "confianza": "ALTA",
                "observacion": "Resolución principal compartida — se copiará a cada carpeta de folio"
            })
        elif any(ig in stem for ig in ["2379", "TERMINO CONVENIO"]):
            folio_destino = "77693" if "2379" in stem else "RADIO AMIGA"
            filas.append({
                "tipo": "RESOLUCION", "archivo_origen": ruta,
                "folio": folio_destino, "razon_social": "FUERA DE LIBRO1 o terminado",
                "nombre_destino": "N/A", "confianza": "IGNORAR",
                "observacion": "No corresponde a ningún folio activo en Libro1"
            })
        else:
            filas.append({
                "tipo": "RESOLUCION", "archivo_origen": ruta,
                "folio": "REFERENCIA", "razon_social": "",
                "nombre_destino": "N/A", "confianza": "REVISAR",
                "observacion": "Documento de referencia — evaluar si aplica a algún folio específico"
            })
    return filas


def procesar_colaboradores(folios_libro1):
    """Cruza certificados por RUT con Libro1."""
    filas = []
    carpeta = ARCHIVOS / "Registro de Colaboradores del Estado - Registros 19862"

    # Construir índice RUT → folio desde Libro1
    rut_a_folio = {}
    for folio, info in folios_libro1.items():
        rut_raw = info["rut"]
        # Normalizar RUT: quitar puntos, guion, espacio
        rut_norm = re.sub(r'[.\-\s]', '', rut_raw).upper()
        rut_a_folio[rut_norm] = folio

    encontrados = 0
    no_encontrados = 0
    for f in sorted(os.listdir(carpeta)):
        if not f.endswith(".pdf"):
            continue
        m = re.match(r'certificado_(\d+)_([0-9Kk])\.pdf', f)
        if not m:
            continue
        rut_num = m.group(1)
        dv = m.group(2).upper()
        rut_norm = rut_num + dv

        folio = rut_a_folio.get(rut_norm)
        if folio:
            encontrados += 1
            info = folios_libro1[folio]
            filas.append({
                "tipo": "COLABORADOR", "archivo_origen": str(carpeta / f),
                "folio": folio, "razon_social": info["nombre"],
                "nombre_destino": f"(permanece en carpeta Colaboradores)",
                "confianza": "ALTA",
                "observacion": f"RUT {rut_num}-{dv} → folio {folio}"
            })
        else:
            no_encontrados += 1

    print(f"Colaboradores: {encontrados} cruzan con Libro1, {no_encontrados} sin coincidencia")
    # Solo mostrar los que cruzan (los demás son de otros programas/fondos)
    return filas


# ── generar Excel de reporte ──────────────────────────────────────────────────

COLOR = {
    "ALTA":    "C6EFCE",  # verde
    "REVISAR": "FFEB9C",  # amarillo
    "MANUAL":  "FFC7CE",  # rojo claro
    "IGNORAR": "D9D9D9",  # gris
}

def escribir_excel(filas):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapeo Propuesto"

    headers = ["TIPO", "ARCHIVO ORIGEN", "FOLIO", "RAZON SOCIAL", "NOMBRE DESTINO", "CONFIANZA", "OBSERVACIÓN"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for fila in filas:
        row = [
            fila["tipo"], fila["archivo_origen"], str(fila["folio"]),
            fila["razon_social"], fila["nombre_destino"],
            fila["confianza"], fila["observacion"]
        ]
        ws.append(row)
        color = COLOR.get(fila["confianza"], "FFFFFF")
        fill = PatternFill("solid", fgColor=color)
        for cell in ws[ws.max_row]:
            cell.fill = fill

    # Resumen en hoja 2
    ws2 = wb.create_sheet("Resumen")
    ws2.append(["Estado", "Cantidad"])
    conteos = {}
    for f in filas:
        c = f["confianza"]
        conteos[c] = conteos.get(c, 0) + 1
    for k, v in sorted(conteos.items()):
        ws2.append([k, v])

    # Ajustar anchos
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)

    wb.save(OUT)
    print(f"\nReporte guardado en: {OUT}")


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    folios_libro1 = leer_libro1()
    print(f"Folios en Libro1: {len(folios_libro1)}")

    todas_filas = []
    todas_filas += procesar_convenios(folios_libro1)
    todas_filas += procesar_egresos(folios_libro1)
    todas_filas += procesar_rendiciones(folios_libro1)
    todas_filas += procesar_garantias(folios_libro1)
    todas_filas += procesar_resoluciones(folios_libro1)
    colaboradores = procesar_colaboradores(folios_libro1)

    print(f"\n=== RESUMEN ===")
    por_tipo = {}
    for f in todas_filas:
        t = f["tipo"]
        por_tipo[t] = por_tipo.get(t, 0) + 1
    for t, n in sorted(por_tipo.items()):
        print(f"  {t}: {n} archivos")

    por_confianza = {}
    for f in todas_filas:
        c = f["confianza"]
        por_confianza[c] = por_confianza.get(c, 0) + 1
    print()
    for c, n in sorted(por_confianza.items()):
        print(f"  {c}: {n}")

    print(f"\nCertificados colaboradores que cruzan con Libro1: {len(colaboradores)}")

    escribir_excel(todas_filas)

    # Mostrar casos REVISAR y MANUAL
    print("\n=== CASOS QUE REQUIEREN REVISIÓN MANUAL ===")
    for f in todas_filas:
        if f["confianza"] in ("REVISAR", "MANUAL"):
            print(f"  [{f['confianza']}] {Path(f['archivo_origen']).name} → folio {f['folio']}")
            if f["observacion"]:
                print(f"         → {f['observacion']}")


if __name__ == "__main__":
    main()
