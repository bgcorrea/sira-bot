"""
Genera data/master_subida.xlsx para los 89 folios de Valparaíso 2023
apuntando a ARCHIVOS-NORMALIZADOS/ y certificados por RUT.
"""

import csv, re, openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

BASE     = Path(__file__).parent.parent
NORM     = BASE / "ARCHIVOS-NORMALIZADOS"
COLAB    = BASE / "ARCHIVOS-SUBIR" / "Registro de Colaboradores del Estado - Registros 19862"
LIBRO1   = BASE / "Libro1.xlsx"
OUT_XLSX = BASE / "data" / "master_subida.xlsx"
OUT_CSV  = BASE / "logs" / "master_subida.csv"

CAMPOS = [
    "folio", "ano", "fondo", "region", "alcance",
    "razon_social", "nombre_fantasia", "rut",
    "estado_sira", "cuotas_declaradas", "monto_garantia", "acto_admin_numero",
    "carpeta_folio",
    "convenio_pdf", "acto_admin_pdf", "certificado_pdf",
    "voucher_pdf", "rendicion_pdf", "garantia_pdf",
    "confianza_convenio", "confianza_acto", "confianza_certificado",
    "confianza_voucher", "confianza_rendicion", "confianza_garantia",
    "n_archivos_encontrados", "observaciones", "estado_carga",
]

# ── índice de certificados por RUT ────────────────────────────────────────────

def indice_certificados():
    idx = {}
    for f in COLAB.glob("certificado_*.pdf"):
        m = re.match(r"certificado_(\d+)_([0-9Kk])\.pdf", f.name)
        if m:
            key = m.group(1) + m.group(2).upper()
            idx[key] = f
    return idx


def buscar_certificado(rut_raw: str, cert_idx: dict):
    rut = re.sub(r"[\.\s]", "", rut_raw)
    rut = rut.replace("-", "")
    if not rut:
        return "", ""
    num = rut[:-1]
    dv  = rut[-1].upper()
    ruta = cert_idx.get(num + dv)
    return (str(ruta), "ALTA") if ruta else ("", "")


# ── encontrar carpeta del folio en ARCHIVOS-NORMALIZADOS ─────────────────────

def carpeta_folio(folio: int) -> Path | None:
    patron = re.compile(rf"^{folio}\s*-")
    for sub in NORM.iterdir():
        if sub.is_dir() and patron.match(sub.name):
            return sub
    return None


def buscar_doc(carpeta: Path, tipo: str, folio: int) -> str:
    nombre = f"{tipo} - {folio}.pdf"
    ruta = carpeta / nombre
    return str(ruta) if ruta.exists() else ""


# ── leer Libro1 ───────────────────────────────────────────────────────────────

def leer_libro1():
    wb = openpyxl.load_workbook(LIBRO1)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        if not row[8]: continue
        rows.append({
            "fondo":   str(row[0]) if row[0] else "",
            "ano":     int(row[1]) if row[1] else 2023,
            "region":  str(row[3]) if row[3] else "VALPARAÍSO",
            "estado":  str(row[4]) if row[4] else "",
            "rut":     str(row[5]) if row[5] else "",
            "folio":   int(row[8]),
            "acto":    str(row[9]) if row[9] else "",
            "nombre":  str(row[11]) if row[11] else "",
            "alcance": str(row[12]) if row[12] else "",
            "monto":   str(row[14]) if row[14] else "",
            "cuotas":  str(row[15]) if row[15] else "1",
        })
    return rows


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    cert_idx = indice_certificados()
    libro1   = leer_libro1()

    filas = []
    for info in libro1:
        folio  = info["folio"]
        carpeta = carpeta_folio(folio)

        convenio  = buscar_doc(carpeta, "Convenio",  folio) if carpeta else ""
        resolucion= buscar_doc(carpeta, "Resolución", folio) if carpeta else ""
        egreso    = buscar_doc(carpeta, "Egreso",    folio) if carpeta else ""
        rendicion = buscar_doc(carpeta, "Rendición", folio) if carpeta else ""
        garantia  = buscar_doc(carpeta, "Garantía",  folio) if carpeta else ""
        cert, conf_cert = buscar_certificado(info["rut"], cert_idx)

        def conf(path): return "ALTA" if path else ""

        obs = []
        if not convenio:   obs.append("Sin convenio")
        if not resolucion: obs.append("Sin resolución")
        if not cert:       obs.append("Sin certificado")
        if not egreso:     obs.append("Sin egreso")
        if not rendicion:  obs.append("Sin rendición")

        n = sum(bool(p) for p in [convenio, resolucion, cert, egreso, rendicion, garantia])

        filas.append({
            "folio":               str(folio),
            "ano":                 info["ano"],
            "fondo":               info["fondo"],
            "region":              info["region"],
            "alcance":             info["alcance"],
            "razon_social":        info["nombre"],
            "nombre_fantasia":     info["nombre"],
            "rut":                 info["rut"],
            "estado_sira":         "Borrador",
            "cuotas_declaradas":   info["cuotas"],
            "monto_garantia":      info["monto"],
            "acto_admin_numero":   info["acto"],
            "carpeta_folio":       str(carpeta) if carpeta else "",
            "convenio_pdf":        convenio,
            "acto_admin_pdf":      resolucion,
            "certificado_pdf":     cert,
            "voucher_pdf":         egreso,
            "rendicion_pdf":       rendicion,
            "garantia_pdf":        garantia,
            "confianza_convenio":  conf(convenio),
            "confianza_acto":      conf(resolucion),
            "confianza_certificado": conf_cert,
            "confianza_voucher":   conf(egreso),
            "confianza_rendicion": conf(rendicion),
            "confianza_garantia":  conf(garantia),
            "n_archivos_encontrados": n,
            "observaciones":       " | ".join(obs) if obs else "OK",
            "estado_carga":        "",
        })

    # ── escribir Excel ────────────────────────────────────────────────────────
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "master_subida"

    HDR_FILL = PatternFill("solid", fgColor="1F4E78")
    HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
    ALTA_FILL = PatternFill("solid", fgColor="C6EFCE")
    VACIO_FILL = PatternFill("solid", fgColor="FFC7CE")

    for ci, campo in enumerate(CAMPOS, 1):
        c = ws_out.cell(row=1, column=ci, value=campo)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center")

    CONF_PATH = [
        ("confianza_convenio",    "convenio_pdf"),
        ("confianza_acto",        "acto_admin_pdf"),
        ("confianza_certificado", "certificado_pdf"),
        ("confianza_voucher",     "voucher_pdf"),
        ("confianza_rendicion",   "rendicion_pdf"),
        ("confianza_garantia",    "garantia_pdf"),
    ]

    for ri, fila in enumerate(filas, 2):
        for ci, campo in enumerate(CAMPOS, 1):
            ws_out.cell(row=ri, column=ci, value=fila.get(campo, ""))
        for conf_campo, path_campo in CONF_PATH:
            ci_conf = CAMPOS.index(conf_campo) + 1
            ci_path = CAMPOS.index(path_campo) + 1
            fill = ALTA_FILL if fila.get(conf_campo) == "ALTA" else (VACIO_FILL if not fila.get(path_campo) else None)
            if fill:
                ws_out.cell(row=ri, column=ci_conf).fill = fill
                ws_out.cell(row=ri, column=ci_path).fill = fill

    anchos = {"folio":8,"ano":6,"fondo":8,"region":14,"alcance":12,
              "razon_social":45,"nombre_fantasia":30,"rut":14,
              "estado_sira":11,"cuotas_declaradas":8,"monto_garantia":14,
              "acto_admin_numero":14,"carpeta_folio":70,
              "convenio_pdf":70,"acto_admin_pdf":70,"certificado_pdf":70,
              "voucher_pdf":70,"rendicion_pdf":70,"garantia_pdf":70}
    for ci, campo in enumerate(CAMPOS, 1):
        ws_out.column_dimensions[get_column_letter(ci)].width = anchos.get(campo, 12)
    ws_out.freeze_panes = "A2"

    OUT_XLSX.parent.mkdir(exist_ok=True)
    wb_out.save(OUT_XLSX)

    # ── CSV ───────────────────────────────────────────────────────────────────
    with open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CAMPOS, extrasaction="ignore")
        w.writeheader(); w.writerows(filas)

    # ── resumen ───────────────────────────────────────────────────────────────
    completos = sum(1 for f in filas if f["observaciones"] == "OK")
    sin_cert  = sum(1 for f in filas if not f["certificado_pdf"])
    print(f"Master generado: {OUT_XLSX}")
    print(f"  Total folios   : {len(filas)}")
    print(f"  Completos (OK) : {completos}")
    print(f"  Sin certificado: {sin_cert}  ← personas naturales")
    print(f"  Otros faltantes: {len(filas) - completos - sin_cert}")
    print(f"\nPróximo paso: python scripts/04_subir_documentos.py")


if __name__ == "__main__":
    main()
