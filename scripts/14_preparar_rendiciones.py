"""
Script 14 — Preparar rendiciones faltantes para subida.
=========================================================
Lee los PDFs en rendiciones_faltantes/ (nombrados {folio}.pdf),
verifica que cada folio esté en master_subida.xlsx, y actualiza
la columna rendicion_pdf con la ruta absoluta al PDF.

Uso:
    python scripts/14_preparar_rendiciones.py            # dry-run (solo reporta)
    python scripts/14_preparar_rendiciones.py --aplicar  # actualiza master
"""

import argparse
from pathlib import Path

import openpyxl

RENDICIONES_DIR = Path("/home/bgcorrea/personal/workspace/caigg/rendiciones_faltantes")
MASTER          = Path("data/master_subida.xlsx")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--aplicar", action="store_true",
                        help="Actualizar master_subida.xlsx con las rutas")
    args = parser.parse_args()

    if not RENDICIONES_DIR.exists():
        print(f"[ERROR] No existe: {RENDICIONES_DIR}")
        return
    if not MASTER.exists():
        print(f"[ERROR] No existe: {MASTER}")
        return

    # Leer PDFs disponibles: {folio -> ruta}
    pdfs = {f.stem: f for f in sorted(RENDICIONES_DIR.iterdir())
            if f.is_file() and f.suffix.lower() == ".pdf"}
    print(f"\nPDFs encontrados en rendiciones_faltantes/: {len(pdfs)}")

    # Leer master
    wb = openpyxl.load_workbook(MASTER)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    col_folio = headers.index("folio") + 1
    col_rend  = headers.index("rendicion_pdf") + 1

    actualizados = []
    no_en_master = []
    ya_tenia     = []

    folio_a_fila = {}
    for row in ws.iter_rows(min_row=2):
        fval = str(row[col_folio - 1].value).strip() if row[col_folio - 1].value else ""
        if fval:
            folio_a_fila[fval] = row

    for folio, pdf_path in sorted(pdfs.items()):
        if folio not in folio_a_fila:
            no_en_master.append(folio)
            continue
        row = folio_a_fila[folio]
        val_actual = str(row[col_rend - 1].value or "").strip()
        if val_actual and val_actual != "nan" and Path(val_actual).exists():
            ya_tenia.append((folio, val_actual))
        else:
            actualizados.append((folio, str(pdf_path.resolve())))
            if args.aplicar:
                row[col_rend - 1].value = str(pdf_path.resolve())

    print(f"\n{'='*65}")
    print(f"  REPORTE — RENDICIONES A PREPARAR")
    print(f"{'='*65}")
    print(f"\n  Listos para subir (rendicion_pdf vacío → se asignará): {len(actualizados)}")
    for folio, ruta in actualizados:
        print(f"    {folio}  →  {ruta}")

    if ya_tenia:
        print(f"\n  Ya tenían rendicion_pdf asignado: {len(ya_tenia)}")
        for folio, ruta in ya_tenia:
            print(f"    {folio}  →  {ruta}")

    if no_en_master:
        print(f"\n  [WARN] PDFs sin folio en master: {no_en_master}")

    if args.aplicar:
        wb.save(MASTER)
        print(f"\n✅  Master actualizado: {len(actualizados)} folios con rendicion_pdf → {MASTER}")
        print()
        print("  Próximo paso:")
        print("  python scripts/04_subir_documentos.py --solo-rendicion")
    else:
        print(f"\n  → Para aplicar:")
        print(f"    python scripts/14_preparar_rendiciones.py --aplicar")
    print()


if __name__ == "__main__":
    main()
