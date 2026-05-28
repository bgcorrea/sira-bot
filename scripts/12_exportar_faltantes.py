"""
Script 12 — Exportar Excel de documentos faltantes para gestión manual.

Genera logs/documentos_faltantes.xlsx con:
  - Hoja 1 "LISTOS": folios que pueden enviarse hoy
  - Hoja 2 "FALTANTES": folios incompletos con columnas de checkbox por documento

Uso:
    python scripts/12_exportar_faltantes.py
"""

from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

MASTER       = Path("data/master_subida.xlsx")
DISTRIBUCION = Path("data/DISTRIBUCIÓN CARGA VB.xlsx")
OUT          = Path("logs/documentos_faltantes.xlsx")

CAMPOS = {
    "convenio_pdf":    "Convenio",
    "acto_admin_pdf":  "Resolución",
    "certificado_pdf": "Certificado",
    "voucher_pdf":     "Transferencias",
    "rendicion_pdf":   "Rendición",
}

# Colores
VERDE       = "C6EFCE"
ROJO        = "FFC7CE"
AMARILLO    = "FFEB9C"
GRIS_HEADER = "2F5496"
AZUL_LISTO  = "D9E1F2"
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN        = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _set_col_width(ws, col_letter: str, width: float):
    ws.column_dimensions[col_letter].width = width


def _header_row(ws, valores: list[str], fila: int = 1):
    fill = PatternFill("solid", fgColor=GRIS_HEADER)
    for col, val in enumerate(valores, 1):
        cell = ws.cell(row=fila, column=col, value=val)
        cell.font     = HEADER_FONT
        cell.fill     = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border   = THIN


def _tiene_archivo(val) -> bool:
    v = str(val).strip()
    return bool(v) and v != "nan" and Path(v).exists()


def build_faltantes_data(master: pd.DataFrame, folios_dist: set) -> tuple[list, list]:
    listos     = []
    faltantes  = []

    for _, row in master.iterrows():
        folio = str(row["folio"])
        if folio not in folios_dist:
            continue

        region = str(row.get("region", "")).strip()
        razon  = str(row.get("razon_social", "")).strip()
        rut    = str(row.get("rut", "")).strip()

        campos_faltantes = {}
        for campo, nombre in CAMPOS.items():
            campos_faltantes[nombre] = not _tiene_archivo(row.get(campo, ""))

        faltan = [n for n, falta in campos_faltantes.items() if falta]

        base = {
            "Folio":        folio,
            "Región":       region,
            "Organización": razon,
            "RUT":          rut if rut and rut != "nan" else "",
        }

        if not faltan:
            listos.append(base)
        else:
            faltantes.append({
                **base,
                "Documentos faltantes": "; ".join(faltan),
                **{f"✓ {n}": ("FALTA" if v else "") for n, v in campos_faltantes.items()},
            })

    listos.sort(key=lambda r: (r["Región"], r["Folio"]))
    faltantes.sort(key=lambda r: (r["Región"], r["Folio"]))
    return listos, faltantes


def escribir_hoja_listos(wb, listos: list):
    ws = wb.active
    ws.title = f"LISTOS ({len(listos)})"

    cols = ["Folio", "Región", "Organización", "RUT"]
    _header_row(ws, cols)
    _set_col_width(ws, "A", 10)
    _set_col_width(ws, "B", 14)
    _set_col_width(ws, "C", 55)
    _set_col_width(ws, "D", 16)
    ws.row_dimensions[1].height = 22

    fill_listo = PatternFill("solid", fgColor=VERDE)
    for i, row in enumerate(listos, 2):
        for j, col in enumerate(cols, 1):
            cell = ws.cell(row=i, column=j, value=row[col])
            cell.fill   = fill_listo
            cell.border = THIN
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def escribir_hoja_faltantes(wb, faltantes: list):
    ws = wb.create_sheet(title=f"FALTANTES ({len(faltantes)})")

    doc_cols = [f"✓ {n}" for n in CAMPOS.values()]
    cols = ["Folio", "Región", "Organización", "RUT", "Documentos faltantes"] + doc_cols
    _header_row(ws, cols)

    anchos = [10, 14, 55, 16, 45] + [15] * len(doc_cols)
    for i, w in enumerate(anchos, 1):
        _set_col_width(ws, get_column_letter(i), w)
    ws.row_dimensions[1].height = 22

    fill_ok    = PatternFill("solid", fgColor=VERDE)
    fill_falta = PatternFill("solid", fgColor=ROJO)
    fill_base  = PatternFill("solid", fgColor="FFFFFF")

    for i, row in enumerate(faltantes, 2):
        for j, col in enumerate(cols, 1):
            val  = row.get(col, "")
            cell = ws.cell(row=i, column=j, value=val)
            cell.border    = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=(col == "Organización"))

            if col in doc_cols:
                cell.fill      = fill_falta if val == "FALTA" else fill_ok
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.value     = "✗ FALTA" if val == "FALTA" else "✓"
            else:
                cell.fill = fill_base

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def main() -> None:
    master = pd.read_excel(MASTER)
    dist   = pd.read_excel(DISTRIBUCION)
    folios_dist = set(dist["ID Convenio"].astype(str).str.strip())

    listos, faltantes = build_faltantes_data(master, folios_dist)

    wb = openpyxl.Workbook()
    escribir_hoja_listos(wb, listos)
    escribir_hoja_faltantes(wb, faltantes)

    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)

    print(f"\n{'='*55}")
    print(f"  EXCEL DE GESTIÓN EXPORTADO")
    print(f"{'='*55}")
    print(f"\n  Hoja 1 — LISTOS:     {len(listos):>3} folios")
    print(f"  Hoja 2 — FALTANTES:  {len(faltantes):>3} folios")
    print(f"\n  → {OUT}")
    print()

    # Mini-resumen de faltantes por región y documento
    print("  ── Qué buscar por región ──")
    from collections import Counter
    por_region: dict[str, Counter] = {}
    for r in faltantes:
        reg = r["Región"]
        por_region.setdefault(reg, Counter())
        for nombre in CAMPOS.values():
            if r.get(f"✓ {nombre}") == "FALTA":
                por_region[reg][nombre] += 1

    for reg in sorted(por_region):
        total = sum(1 for r in faltantes if r["Región"] == reg)
        items = ", ".join(f"{n} ({c})" for n, c in por_region[reg].most_common())
        print(f"  {reg:<15} {total:>3} folios → {items}")
    print()


if __name__ == "__main__":
    main()
